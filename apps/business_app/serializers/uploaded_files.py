import os
from io import BytesIO
import logging
import json

import pandas as pd
from django.core.files.base import ContentFile
from openpyxl import load_workbook
from rest_framework import serializers

from apps.business_app.models import UploadedFiles
from apps.business_app.models import StudyType

from apps.business_app.serializers.allele_nodes import AlleleNodeSerializer
from apps.business_app.serializers.pdb_files import PdbFilesSerializer
from apps.business_app.serializers.study import StudySerializerShort


logger = logging.getLogger(__name__)


class UploadedFileToCompareVsStudiesSerializer(serializers.Serializer):
    file = serializers.FileField(required=True)

    def validate_file(self, value):
        study_types_sheet = StudyType.objects.only("sheet_name").values_list(
            "sheet_name", flat=True
        )
        try:
            excel_file = pd.ExcelFile(value)
            file_sheets = set(excel_file.sheet_names)
            study_types_set = set(study_types_sheet)

            unmatched_sheets_on_excel = (
                file_sheets
                - study_types_set
                - UploadedFiles.SHEETS_TO_OMMIT_IN_PROCESSING
            )
            unmatched_sheets_on_studies = study_types_set - file_sheets

            if unmatched_sheets_on_excel or unmatched_sheets_on_studies:
                raise serializers.ValidationError(
                    {
                        "unmatched_sheets_on_excel": list(unmatched_sheets_on_excel),
                        "unmatched_sheets_on_studies": list(
                            unmatched_sheets_on_studies
                        ),
                    }
                )

            return value
        except Exception as e:
            if isinstance(e, serializers.ValidationError):
                raise
            raise serializers.ValidationError(f"Error al validar el archivo: {str(e)}")


class SimpleListUploadedFilesSerializer(serializers.ModelSerializer):
    gene_name = serializers.CharField(source="gene.name", read_only=True, default=None)
    studies = StudySerializerShort(many=True, read_only=True)

    class Meta:
        model = UploadedFiles
        fields = [
            "id",
            "custom_name",
            "description",
            "original_file",
            "system_user",
            "gene",
            "gene_name",
            "predefined",
            "studies",
        ]
        read_only_fields = [
            "id",
        ]

    def save(self):
        try:
            return super().save()
        except Exception as e:
            logger.exception(f"{str(e)}")
            raise serializers.ValidationError(e) from e


class SheetStudyAssignmentSerializer(serializers.Serializer):
    sheet_name = serializers.CharField()
    study_type = serializers.PrimaryKeyRelatedField(
        queryset=StudyType.objects.all(), allow_null=True, required=False
    )


class UploadedFilesSerializer(SimpleListUploadedFilesSerializer):
    pdb_files = PdbFilesSerializer(many=True, read_only=True)
    allele_nodes = AlleleNodeSerializer(many=True, read_only=True)
    sheet_study_assignments = serializers.CharField(
        write_only=True, required=False, allow_blank=True
    )

    class Meta(SimpleListUploadedFilesSerializer.Meta):
        fields = SimpleListUploadedFilesSerializer.Meta.fields + [
            "pdb_files",
            "allele_nodes",
            "sheet_study_assignments",
        ]
        read_only_fields = SimpleListUploadedFilesSerializer.Meta.read_only_fields + [
            "gene_name",
            "predefined",
            "pdb_files",
            "allele_nodes",
        ]

    def _rename_excel_sheets(self, original_file, rename_pairs):
        """Return a new in-memory Excel file with renamed worksheets."""
        original_file.seek(0)
        workbook = load_workbook(filename=BytesIO(original_file.read()))

        for source_name, target_name in rename_pairs:
            if source_name not in workbook.sheetnames:
                raise serializers.ValidationError(
                    {
                        "sheet_study_assignments": [
                            f"La pesta\u00f1a '{source_name}' no existe en el archivo cargado."
                        ]
                    }
                )

            if target_name in workbook.sheetnames and target_name != source_name:
                raise serializers.ValidationError(
                    {
                        "sheet_study_assignments": [
                            f"No se puede renombrar '{source_name}' a '{target_name}' porque ya existe una pesta\u00f1a con ese nombre."
                        ]
                    }
                )

            workbook[source_name].title = target_name

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return ContentFile(
            output.getvalue(),
            name=os.path.basename("fixed" + original_file.name),
        )

    def _build_sheet_rename_pairs(self, sheet_study_assignments):
        """Create rename pairs from assignments where study_type is provided."""
        rename_pairs = []
        sheet_study_assignments = sheet_study_assignments.replace("null", "None")
        sheet_study_assignments = SheetStudyAssignmentSerializer(
            data=eval(sheet_study_assignments),
            many=True,
        )
        sheet_study_assignments.is_valid(raise_exception=True)
        sheet_study_assignments = sheet_study_assignments.validated_data
        for assignment in sheet_study_assignments:
            study_type = assignment.get("study_type")
            source_name = assignment.get("sheet_name")
            if not study_type:
                continue

            target_name = study_type.sheet_name
            if source_name != target_name:
                rename_pairs.append((source_name, target_name))

        return rename_pairs

    def validate(self, attrs):
        """Rename worksheet names in original_file according to sheet_study_assignments."""
        attrs = super().validate(attrs)
        sheet_study_assignments = attrs.pop("sheet_study_assignments", [])
        if not sheet_study_assignments:
            return attrs

        original_file = attrs.get("original_file")
        if not original_file:
            raise serializers.ValidationError(
                {
                    "original_file": "Debe enviar original_file cuando usa sheet_study_assignments."
                }
            )

        rename_pairs = self._build_sheet_rename_pairs(sheet_study_assignments)
        if not rename_pairs:
            return attrs

        try:
            attrs["original_file"] = self._rename_excel_sheets(
                original_file, rename_pairs
            )
        except serializers.ValidationError:
            raise
        except Exception as exc:
            raise serializers.ValidationError(
                {"original_file": f"Error al renombrar pesta\u00f1as del Excel: {exc}"}
            )

        return attrs
